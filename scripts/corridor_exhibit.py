#!/usr/bin/env python3
"""One Loop: WS-B2 corridor exhibit.

Produces results/corridor.json with the committed result envelope.

Pre-registered design (committed by this script BEFORE results are read;
seed=42):
  * Target: monthly international visitor arrivals to Singapore by origin
    country (SingStat IVA by place of residence, data.gov.sg dataset
    d_7e7b2ee60c6ffc962f80fef129cf306e), for the top ~12 origin COUNTRIES
    ranked by total 2025 volume; a stated PROXY for cross-border card-spend
    corridors.
  * COVID handling (pre-registered): train window 2008Jan-2024Dec (includes
    the 2020-2022 structural break; noted explicitly in 'covid_note'),
    HOLDOUT = 2025Jan-2026Jan (13 months, post-recovery). Rolling
    one-month-ahead backtest: features at month t use actuals only through
    t-1 (lags) or static covariates; the seasonal-naive baseline likewise
    uses the actual at t-12.
  * Features: target lags (1,2,3,12); rolling means of the previous 3 and 12
    months (windows end at t-1; never same-period); month-of-year
    (categorical); static gravity covariates: KNOMAD-2021 SG-outbound
    bilateral remittance (proxy for origin-country migrant stock resident in
    SG; construction disclosed in 'covariate_provenance'), great-circle
    distance SG->origin capital, shared-official-language flag
    (English/Malay/Mandarin/Tamil). Static or lagged only; NO same-period
    covariates.
  * Model: ONE global LightGBM across corridors (corridor id categorical)
    vs per-corridor SEASONAL NAIVE (t-12). Fixed 500 trees, no early
    stopping (no tuning choices post-registration). deterministic=True,
    n_jobs=1, CPU.
  * Metrics: MAE and MASE per corridor and pooled, on the holdout, reported
    as obtained. MASE scale = in-sample mean |y_t - y_{t-12}| over the train
    window (COVID months included in the scale; disclosed).
  * Grouped reconciliation: country base forecasts -> region -> total
    (hierarchy built over the SELECTED corridors, so it is coherent by
    construction once reconciled). MinT-shrink: Schafer-Strimmer shrunk
    covariance of in-sample base-forecast errors + summing matrix. Base vs
    reconciled holdout accuracy REPORTED, whichever direction.
  * Attributions: LightGBM SHAP (TreeExplainer) mean |value| over holdout
    rows, top-3 features per corridor; labeled MODEL ATTRIBUTIONS, never
    "drivers".

Honesty rider: results reported as obtained, whichever direction.

Usage:
  python scripts/corridor_exhibit.py            # compute + write results/corridor.json
  python scripts/corridor_exhibit.py --check    # recompute (CPU, single-thread,
                                                # deterministic) and verify the
                                                # existing JSON matches at 1e-6
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import sys
from pathlib import Path

import numpy as np

SEED = 42
REPO = Path(__file__).resolve().parent.parent
DATA_DIR = REPO / "data"
OUT_PATH = REPO / "results" / "corridor.json"

SINGSTAT_CSV = DATA_DIR / "singstat_iva_monthly.csv"
KNOMAD_XLSX = DATA_DIR / "knomad_bilateral_2021.xlsx"

N_CORRIDORS = 12
TRAIN_START = (2008, 1)
FEATURE_START = (2009, 1)  # first month with a valid lag-12
TRAIN_END = (2024, 12)
HOLDOUT_START = (2025, 1)
HOLDOUT_END = (2026, 1)

MONTH_ABBR = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}

# --- static gravity covariates (hardcoded, disclosed) -----------------------
SG_LATLON = (1.3521, 103.8198)
CAPITALS = {  # origin -> (lat, lon) of capital / seat of government
    "China": (39.9042, 116.4074), "Indonesia": (-6.2088, 106.8456),
    "Malaysia": (3.1390, 101.6869), "Australia": (-35.2809, 149.1300),
    "India": (28.6139, 77.2090), "Philippines": (14.5995, 120.9842),
    "USA": (38.9072, -77.0369), "Japan": (35.6762, 139.6503),
    "United Kingdom": (51.5074, -0.1278), "South Korea": (37.5665, 126.9780),
    "Taiwan": (25.0330, 121.5654), "Thailand": (13.7563, 100.5018),
    "Vietnam": (21.0285, 105.8542), "Germany": (52.5200, 13.4050),
    "Hong Kong SAR": (22.3193, 114.1694), "Canada": (45.4215, -75.6972),
    "France": (48.8566, 2.3522),
}
# shares an official language with Singapore (English / Malay / Mandarin / Tamil)
COMMON_LANGUAGE = {
    "China": 1, "Indonesia": 1, "Malaysia": 1, "Australia": 1, "India": 1,
    "Philippines": 1, "USA": 1, "Japan": 0, "United Kingdom": 1,
    "South Korea": 0, "Taiwan": 1, "Thailand": 0, "Vietnam": 0,
    "Germany": 0, "Hong Kong SAR": 1, "Canada": 1, "France": 0,
}
# SingStat origin name -> KNOMAD receiving-country column name (None = absent)
KNOMAD_NAME = {
    "China": "China", "Indonesia": "Indonesia", "Malaysia": "Malaysia",
    "Australia": "Australia", "India": "India", "Philippines": "Philippines",
    "USA": "United States", "Japan": "Japan",
    "United Kingdom": "United Kingdom", "South Korea": "Korea, Rep.",
    "Taiwan": None, "Thailand": "Thailand", "Vietnam": "Vietnam",
    "Germany": "Germany", "Hong Kong SAR": "Hong Kong SAR, China",
    "Canada": "Canada", "France": "France",
}

FEATURES = ["corridor_id", "month", "lag1", "lag2", "lag3", "lag12",
            "roll3", "roll12", "knomad_remit_log", "dist_log", "common_language"]
CATEGORICAL = ["corridor_id", "month"]


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def verify_checksums() -> dict[str, str]:
    """Compute sha256 of inputs and cross-check against data/CHECKSUMS.txt."""
    pinned = {}
    with open(DATA_DIR / "CHECKSUMS.txt") as f:
        for line in f:
            parts = line.split()
            if len(parts) == 2:
                pinned[Path(parts[1]).name] = parts[0]
    out = {}
    for path in (SINGSTAT_CSV, KNOMAD_XLSX):
        digest = sha256_of(path)
        want = pinned.get(path.name)
        if want is not None and want != digest:
            raise SystemExit(f"CHECKSUM MISMATCH for {path.name}: "
                             f"pinned {want} != computed {digest}")
        out[path.name] = digest
    return out


def parse_month(label: str) -> tuple[int, int]:
    return int(label[:4]), MONTH_ABBR[label[4:]]


def month_leq(a, b):  # (y,m) tuples
    return a <= b


def parse_singstat():
    """Parse the wide SingStat CSV. Leading spaces encode hierarchy:
    0 = total row, 4 = region, 8 = country. Returns months ascending plus
    {country: (region, np.array arrivals ascending)}."""
    with open(SINGSTAT_CSV, newline="") as f:
        rows = list(csv.reader(f))
    header = rows[0]
    month_labels = header[1:]
    months = [parse_month(m) for m in month_labels]
    order = np.argsort([y * 100 + m for y, m in months])  # ascending
    months_asc = [months[i] for i in order]

    countries = {}
    current_region = None
    for row in rows[1:]:
        name = row[0]
        indent = len(name) - len(name.lstrip(" "))
        clean = name.strip()
        vals = np.array(
            [float(v) if v not in ("", "na", "-") else np.nan for v in row[1:]],
            dtype=np.float64)[order]
        if indent == 0:
            continue  # grand-total row: not used (our total = sum of selected)
        if indent == 4:
            current_region = clean
        elif indent == 8:
            if clean.startswith("Other Markets"):
                continue
            if current_region is None:
                raise SystemExit(f"country row before any region row: {clean}")
            countries[clean] = (current_region, vals)
        else:
            raise SystemExit(f"unexpected indent {indent} for row {clean!r}")
    return months_asc, countries


def load_knomad_sg_row() -> dict[str, float]:
    """KNOMAD bilateral remittance matrix 2021: row = sending country,
    column = receiving country. The Singapore SENDING row (USD millions,
    SG -> origin) proxies the origin-country migrant stock resident in SG
    (migrants remit home). Returns receiving-country -> value."""
    import openpyxl

    wb = openpyxl.load_workbook(KNOMAD_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    data = [list(r) for r in ws.iter_rows(values_only=True)]
    recv_header = data[1]
    sg_row = None
    for r in data[2:]:
        if r[0] and str(r[0]).strip().lower() == "singapore":
            sg_row = r
            break
    if sg_row is None:
        raise SystemExit("Singapore sending row not found in KNOMAD matrix")
    out = {}
    for j in range(1, len(recv_header)):
        name = recv_header[j]
        if name is None:
            continue
        v = sg_row[j]
        if v is not None:
            out[str(name).strip()] = float(v)
    return out


def haversine_km(a, b):
    lat1, lon1, lat2, lon2 = map(math.radians, (*a, *b))
    dlat, dlon = lat2 - lat1, lon2 - lon1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * 6371.0 * math.asin(math.sqrt(h))


def build_series(months, countries):
    """Select top-N corridors by 2025 volume; build node dict:
    total -> regions(selected members only) -> countries. Returns
    (node_names ordered [total, regions..., countries...], series matrix,
    bottom country list, region map)."""
    idx_2025 = [i for i, (y, _m) in enumerate(months) if y == 2025]
    vol = {c: np.nansum(v[idx_2025]) for c, (_r, v) in countries.items()}
    top = sorted(vol, key=lambda c: -vol[c])[:N_CORRIDORS]
    top = sorted(top, key=lambda c: -vol[c])  # deterministic order by volume

    region_of = {c: countries[c][0] for c in top}
    regions = sorted({region_of[c] for c in top})
    T = len(months)
    series = {}
    for c in top:
        series[c] = countries[c][1]
        if np.isnan(series[c]).any():
            raise SystemExit(f"NaNs in series for {c}")
    for r in regions:
        members = [c for c in top if region_of[c] == r]
        series[f"REGION::{r}"] = np.sum([series[c] for c in members], axis=0)
    series["TOTAL::selected"] = np.sum([series[c] for c in top], axis=0)

    node_names = (["TOTAL::selected"] + [f"REGION::{r}" for r in regions] + top)
    mat = np.stack([series[n] for n in node_names])  # (n_nodes, T)
    return node_names, mat, top, region_of, regions, T


def make_features(node_names, mat, months, knomad_sg, top):
    """Long-format feature table for every node x month with valid lag-12."""
    n_nodes, T = mat.shape
    month_num = np.array([m for (_y, m) in months])
    static = {}
    for i, node in enumerate(node_names):
        if node in top:
            kn = KNOMAD_NAME.get(node)
            remit = knomad_sg.get(kn) if kn else None
            static[i] = (
                math.log1p(remit) if remit is not None else np.nan,
                math.log(haversine_km(SG_LATLON, CAPITALS[node])),
                float(COMMON_LANGUAGE[node]),
            )
        else:  # aggregate nodes: no single-origin static covariates
            static[i] = (np.nan, np.nan, np.nan)

    rows, targets, keys = [], [], []
    for i in range(n_nodes):
        y = mat[i]
        for t in range(12, T):
            lag1, lag2, lag3, lag12 = y[t - 1], y[t - 2], y[t - 3], y[t - 12]
            roll3 = y[t - 3:t].mean()
            roll12 = y[t - 12:t].mean()
            rows.append([i, month_num[t], lag1, lag2, lag3, lag12,
                         roll3, roll12, *static[i]])
            targets.append(math.log1p(y[t]))
            keys.append((i, months[t]))
    X = np.array(rows, dtype=np.float64)
    yv = np.array(targets, dtype=np.float64)
    return X, yv, keys


def schafer_strimmer_cov(resid):
    """Shrunk covariance (Schafer-Strimmer 2005: shrink correlations toward
    identity, i.e. covariance toward its diagonal). resid: (T, n)."""
    T, n = resid.shape
    xc = resid - resid.mean(axis=0)
    sd = xc.std(axis=0, ddof=1)
    sd = np.where(sd == 0, 1.0, sd)
    xs = xc / sd
    S = (xc.T @ xc) / (T - 1)                      # sample covariance
    R = (xs.T @ xs) / (T - 1)                      # sample correlation
    # var of each off-diagonal correlation estimate
    w = xs[:, :, None] * xs[:, None, :]            # (T, n, n)
    wbar = w.mean(axis=0)
    var_r = (T / ((T - 1) ** 3)) * ((w - wbar) ** 2).sum(axis=0)
    off = ~np.eye(n, dtype=bool)
    denom = (R[off] ** 2).sum()
    lam = 1.0 if denom == 0 else float(np.clip(var_r[off].sum() / denom, 0.0, 1.0))
    W = lam * np.diag(np.diag(S)) + (1 - lam) * S
    return W, lam


def mint_reconcile(S_mat, W, base):  # base: (n_nodes, h)
    """MinT: reconciled = S (S' W^-1 S)^-1 S' W^-1 yhat."""
    Winv = np.linalg.inv(W)
    G = np.linalg.solve(S_mat.T @ Winv @ S_mat, S_mat.T @ Winv)
    return S_mat @ (G @ base)


def compute():
    checksums = verify_checksums()
    months, countries = parse_singstat()
    node_names, mat, top, region_of, regions, T = build_series(months, countries)
    knomad_sg = load_knomad_sg_row()
    X, yv, keys = make_features(node_names, mat, months, knomad_sg, top)

    is_train = np.array([month_leq(FEATURE_START, m) and month_leq(m, TRAIN_END)
                         for (_i, m) in keys])
    is_hold = np.array([month_leq(HOLDOUT_START, m) and month_leq(m, HOLDOUT_END)
                        for (_i, m) in keys])
    node_idx = np.array([i for (i, _m) in keys])

    import lightgbm as lgb

    model = lgb.LGBMRegressor(
        objective="regression", n_estimators=500, learning_rate=0.05,
        num_leaves=31, min_child_samples=20, subsample=1.0,
        colsample_bytree=1.0, random_state=SEED, deterministic=True,
        force_row_wise=True, n_jobs=1, verbose=-1)
    cat_idx = [FEATURES.index(c) for c in CATEGORICAL]
    model.fit(X[is_train], yv[is_train], categorical_feature=cat_idx)

    pred_log = model.predict(X)
    pred = np.clip(np.expm1(pred_log), 0.0, None)
    actual = np.expm1(yv)

    n_nodes = len(node_names)
    hold_months = sorted({m for ((_i, m), hld) in zip(keys, is_hold) if hld})
    h = len(hold_months)
    m2col = {m: j for j, m in enumerate(hold_months)}

    base_fc = np.full((n_nodes, h), np.nan)
    actual_h = np.full((n_nodes, h), np.nan)
    for r, ((i, m), hld) in enumerate(zip(keys, is_hold)):
        if hld:
            base_fc[i, m2col[m]] = pred[r]
            actual_h[i, m2col[m]] = actual[r]

    # seasonal naive on holdout + in-sample MASE scale, per node
    month_index = {m: t for t, m in enumerate(months)}
    naive_fc = np.full((n_nodes, h), np.nan)
    scale = np.full(n_nodes, np.nan)
    for i in range(n_nodes):
        y = mat[i]
        for m, j in m2col.items():
            naive_fc[i, j] = y[month_index[m] - 12]
        tr = [t for t, m in enumerate(months)
              if month_leq(FEATURE_START, m) and month_leq(m, TRAIN_END)]
        scale[i] = np.mean([abs(y[t] - y[t - 12]) for t in tr])

    def mae(fc, i):
        return float(np.mean(np.abs(fc[i] - actual_h[i])))

    def mase(fc, i):
        return float(mae(fc, i) / scale[i])

    # --- MinT-shrink grouped reconciliation --------------------------------
    n_bottom = len(top)
    S_mat = np.zeros((n_nodes, n_bottom))
    for bi, c in enumerate(top):
        ci = node_names.index(c)
        S_mat[ci, bi] = 1.0
        S_mat[0, bi] = 1.0  # total
        ri = node_names.index(f"REGION::{region_of[c]}")
        S_mat[ri, bi] = 1.0

    # in-sample base errors (train window, one-step-ahead in-sample residuals)
    train_months = sorted({m for ((_i, m), tr) in zip(keys, is_train) if tr})
    tm2col = {m: j for j, m in enumerate(train_months)}
    resid = np.full((len(train_months), n_nodes), np.nan)
    for r, ((i, m), tr) in enumerate(zip(keys, is_train)):
        if tr:
            resid[tm2col[m], i] = actual[r] - pred[r]
    W, shrink_lambda = schafer_strimmer_cov(resid)
    recon_fc = mint_reconcile(S_mat, W, base_fc)
    # coherence-by-construction guard (ML-Verifier b2 finding): a silent clip after MinT
    # would break the coherent-by-construction claim on a future data refresh
    assert (recon_fc >= 0).all(), "MinT produced negative reconciled values; investigate, do not clip"

    # --- SHAP model attributions (holdout rows, bottom corridors) ----------
    import shap

    explainer = shap.TreeExplainer(model)
    sv = explainer.shap_values(X[is_hold])
    hold_nodes = node_idx[is_hold]
    attributions = {}
    for c in top:
        i = node_names.index(c)
        rows = sv[hold_nodes == i]
        mean_abs = np.abs(rows).mean(axis=0)
        order = np.argsort(-mean_abs)[:3]
        attributions[c] = [
            {"feature": FEATURES[j], "mean_abs_shap_log1p_units": round(float(mean_abs[j]), 6)}
            for j in order]

    # --- the same attributions, grouped into the driver families the brief names --------------
    # The top-3 lists above are dominated by lag1 and lag3 on every corridor, so the explanation
    # a partner sees reads as "last month" and "three months ago". The brief asks for "the key
    # drivers of growth, such as seasonality, traveler mix, customer segments, and category
    # shifts". This groups every feature's mean |SHAP| into the four families this model can
    # speak to, so the honest answer is visible: how much of what the model uses is persistence,
    # how much is seasonality, how much is who is travelling. It measures nothing new; it is the
    # same SHAP matrix summed differently, and the label that this is association rather than
    # cause travels with it.
    FAMILY = {
        "lag1": "recent level (persistence)", "lag2": "recent level (persistence)",
        "lag3": "recent level (persistence)", "roll3": "recent level (persistence)",
        "lag12": "seasonality (same month last year)", "roll12": "seasonality (same month last year)",
        "month": "seasonality (calendar month)",
        "corridor_id": "corridor identity (origin-specific level)",
        "knomad_remit_log": "traveler mix (migrant-stock proxy)",
        "dist_log": "traveler mix (distance)",
        "common_language": "traveler mix (shared language)",
    }
    family_names = sorted(set(FAMILY.values()))
    mean_abs_all = np.abs(sv).mean(axis=0)
    total_all = float(mean_abs_all.sum())
    driver_families = []
    for fam in family_names:
        js = [j for j, f in enumerate(FEATURES) if FAMILY[f] == fam]
        s = float(mean_abs_all[js].sum())
        driver_families.append({
            "family": fam,
            "features": [FEATURES[j] for j in js],
            "mean_abs_shap_log1p_units": round(s, 6),
            "share_of_total_attribution": round(s / total_all, 6) if total_all else 0.0,
            "percent_of_total_attribution": round(100.0 * s / total_all, 4) if total_all else 0.0,
        })
    driver_families.sort(key=lambda r: -r["mean_abs_shap_log1p_units"])
    for k, r in enumerate(driver_families, 1):
        r["rank"] = k
    # what the brief names that this model cannot speak to at all
    drivers_not_in_model = [
        "customer segments (no segment field in public arrivals data)",
        "merchant category shifts (arrivals carry no spend category)",
    ]

    # --- assemble metrics ---------------------------------------------------
    def month_label(m):
        return f"{m[0]:04d}-{m[1]:02d}"

    corridors = []
    for c in top:
        i = node_names.index(c)
        corridors.append({
            "origin": c,
            "region": region_of[c],
            "volume_2025": int(np.nansum(
                mat[i][[t for t, m in enumerate(months) if m[0] == 2025]])),
            "mae_model": round(mae(base_fc, i), 2),
            "mae_seasonal_naive": round(mae(naive_fc, i), 2),
            "mae_model_reconciled": round(mae(recon_fc, i), 2),
            "mase_model": round(mase(base_fc, i), 6),
            "mase_seasonal_naive": round(mase(naive_fc, i), 6),
            "mase_model_reconciled": round(mase(recon_fc, i), 6),
            "attribution_top3": attributions[c],
        })

    bottom_idx = [node_names.index(c) for c in top]
    region_idx = [node_names.index(f"REGION::{r}") for r in regions]

    def pooled(fc, idxs):
        return {
            "mase_macro_mean": round(float(np.mean([mase(fc, i) for i in idxs])), 6),
            "mae_mean": round(float(np.mean([mae(fc, i) for i in idxs])), 2),
        }

    reconciliation_detail = {
        "method": "MinT-shrink (Schafer-Strimmer shrunk covariance of in-sample "
                  "one-step base-forecast errors; summing matrix over "
                  "selected-corridor hierarchy country->region->total)",
        "shrinkage_lambda": round(shrink_lambda, 6),
        "hierarchy": {"total": "TOTAL::selected (sum of the 12 modeled corridors)",
                      "regions": regions,
                      "countries": top},
        # seasonal_naive at every level: the same per-node seasonal naive
        # (actual at t-12), aggregated through the identical pooled()/mase()
        # path on the identical holdout; the naive is coherent by construction
        # (its total is the sum of its bottoms), so it is the honest external
        # bar for base AND reconciled at each level (gate fix 2026-08-22).
        "holdout_accuracy": {
            "countries": {"base": pooled(base_fc, bottom_idx),
                          "reconciled": pooled(recon_fc, bottom_idx),
                          "seasonal_naive": pooled(naive_fc, bottom_idx)},
            "regions": {"base": pooled(base_fc, region_idx),
                        "reconciled": pooled(recon_fc, region_idx),
                        "seasonal_naive": pooled(naive_fc, region_idx)},
            "total": {"base": pooled(base_fc, [0]),
                      "reconciled": pooled(recon_fc, [0]),
                      "seasonal_naive": pooled(naive_fc, [0])},
        },
    }

    # --- figures ------------------------------------------------------------
    context_from = (2023, 1)
    ctx_idx = [t for t, m in enumerate(months) if month_leq(context_from, m)]
    fig_months = [month_label(months[t]) for t in ctx_idx]
    figures = {"corridors": []}
    for c in top:
        i = node_names.index(c)
        figures["corridors"].append({
            "origin": c,
            "series": {
                "months": fig_months,
                "actual": [round(float(mat[i][t]), 1) for t in ctx_idx],
                "forecast_model": [round(float(v), 1) for v in base_fc[i]],
                "forecast_seasonal_naive": [round(float(v), 1) for v in naive_fc[i]],
                "forecast_months": [month_label(m) for m in hold_months],
            },
            "attribution_bars": attributions[c],
        })

    figures_schema = (
        "figures.corridors[] = {origin, series:{months:[YYYY-MM context window "
        f"{month_label(context_from)}..{month_label(hold_months[-1])}], "
        "actual:[arrivals aligned to months], forecast_model:[13 one-step-ahead "
        "holdout forecasts aligned to forecast_months], forecast_seasonal_naive:"
        "[same alignment], forecast_months:[YYYY-MM holdout months]}, "
        "attribution_bars:[{feature, mean_abs_shap_log1p_units} top-3, "
        "descending]}. Render: line chart actual vs the two forecast series "
        "(forecasts non-null only over forecast_months); horizontal bar chart "
        "for attribution_bars, labeled 'model attributions (mean |SHAP|, "
        "log1p-arrivals units)'.")

    import lightgbm
    import openpyxl
    import shap as shap_mod
    import sklearn

    result = {
        "seed": SEED,
        "versions": {
            "python": sys.version.split()[0],
            "numpy": np.__version__,
            "lightgbm": lightgbm.__version__,
            "shap": shap_mod.__version__,
            "scikit-learn": sklearn.__version__,
            "openpyxl": openpyxl.__version__,
        },
        "generated_by": "scripts/corridor_exhibit.py --check-able",
        "data_sources": [
            {"name": "SingStat International Visitor Arrivals by place of "
                     "residence, monthly (data.gov.sg, Open Data Licence)",
             "url": "https://data.gov.sg/datasets/d_7e7b2ee60c6ffc962f80fef129cf306e/view",
             "sha256": checksums["singstat_iva_monthly.csv"]},
            {"name": "KNOMAD bilateral remittance matrix 2021 (static gravity "
                     "covariates only)",
             "url": "https://web.archive.org/web/20230424090916/https://knomad.org/sites/default/files/2022-12/bilateral_remittance_matrix_2021_0.xlsx",
             "sha256": checksums["knomad_bilateral_2021.xlsx"]},
        ],
        "labels": ["proxy: tourism arrivals stand in for cross-border card spend"],
        "target": "SingStat IVA monthly by residence (d_7e7b2ee60c6ffc962f80fef129cf306e)",
        "design": {
            "train_window": "2008-01..2024-12 (features valid from 2009-01: lag-12)",
            "holdout": f"{month_label(hold_months[0])}..{month_label(hold_months[-1])} "
                       "(13 months, one-month-ahead rolling backtest; features use "
                       "actuals only through t-1)",
            "corridor_selection": f"top {N_CORRIDORS} origin countries by 2025 volume",
            "model": "one global LightGBM (500 trees, lr 0.05, num_leaves 31, "
                     "corridor id + month categorical) on log1p(arrivals)",
            "baseline": "per-corridor seasonal naive (actual at t-12)",
            "mase_scale": "in-sample mean |y_t - y_{t-12}| over 2009-01..2024-12 "
                          "(COVID months included in the scale; identical scale "
                          "for model and baseline, so the comparison is unaffected)",
        },
        "covid_note": (
            "The training window contains the 2020-2022 COVID structural break "
            "(arrivals collapse to near zero, then a staggered reopening "
            "recovery). Pre-registered handling: keep the break in training "
            "(the model sees the regime through its lag features), hold out "
            "2025-01..2026-01; 13 post-recovery months; and report accuracy "
            "only on the holdout. The MASE denominator also spans the break, "
            "shrinking MASE values for model and baseline alike; model-vs-naive "
            "comparison is scale-invariant."),
        "covariate_provenance": (
            "KNOMAD bilateral remittance matrices are themselves constructed by "
            "a gravity-style allocation (aggregate remittances distributed using "
            "bilateral migrant stocks and income levels), so they are model "
            "output, not direct measurement. Used here as a STATIC cross-"
            "sectional covariate only; the Singapore SENDING row (SG->origin, "
            "USD millions, 2021 vintage) as a proxy for origin-country migrant "
            "stock resident in SG; never as a target. Distance = great-circle "
            "SG->origin capital; common-language = shares an official language "
            "with SG (English/Malay/Mandarin/Tamil). Taiwan and some high-income "
            "receivers are missing from the KNOMAD matrix -> NaN (LightGBM "
            "handles natively)."),
        "covariate_lag": "t-1+",
        "holdout_months": h,
        "holdout_month_labels": [month_label(m) for m in hold_months],
        "mase_model": round(float(np.mean(
            [mase(base_fc, i) for i in bottom_idx])), 6),
        "mase_seasonal_naive": round(float(np.mean(
            [mase(naive_fc, i) for i in bottom_idx])), 6),
        "mase_definition": "macro mean of per-corridor MASE over the 12 corridors",
        "pooled": {
            "mae_model": round(float(np.mean([mae(base_fc, i) for i in bottom_idx])), 2),
            "mae_seasonal_naive": round(float(np.mean([mae(naive_fc, i) for i in bottom_idx])), 2),
            "mase_model_reconciled": round(float(np.mean(
                [mase(recon_fc, i) for i in bottom_idx])), 6),
        },
        "corridors": corridors,
        "reconciliation": "grouped MinT (coherent by construction); "
                          "base-vs-reconciled accuracy compared",
        "reconciliation_detail": reconciliation_detail,
        "driver_families": driver_families,
        "drivers_named_by_the_brief_this_model_cannot_speak_to": drivers_not_in_model,
        "driver_families_note": (
            "The same holdout SHAP matrix as attribution_top3, summed by feature family across all "
            "twelve corridors, so a reader sees what kind of thing the model relies on rather than "
            "which lag. Association, not cause, exactly as attribution_top3 is labeled."
        ),
        "attribution_label": "model attributions (mean |SHAP| on holdout rows, "
                             "log1p-arrivals units): associations learned by "
                             "the model, not causal drivers",
        "figures_schema": figures_schema,
        "figures": figures,
    }
    return result


def numeric_leaves(obj, path=""):
    if isinstance(obj, dict):
        for k, v in obj.items():
            yield from numeric_leaves(v, f"{path}/{k}")
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            yield from numeric_leaves(v, f"{path}/{i}")
    elif isinstance(obj, bool):
        pass
    elif isinstance(obj, (int, float)):
        yield path, float(obj)


def main():
    check = "--check" in sys.argv
    result = compute()
    if check:
        if not OUT_PATH.exists():
            print(f"CHECK FAIL: {OUT_PATH} does not exist"); return 1
        stored = json.loads(OUT_PATH.read_text())
        a = dict(numeric_leaves(stored))
        b = dict(numeric_leaves(result))
        bad = []
        for k in sorted(set(a) | set(b)):
            if k not in a or k not in b:
                bad.append((k, a.get(k), b.get(k))); continue
            if not math.isclose(a[k], b[k], rel_tol=1e-6, abs_tol=1e-6):
                bad.append((k, a[k], b[k]))
        if bad:
            print(f"CHECK FAIL: {len(bad)} numeric leaves differ (>1e-6):")
            for k, va, vb in bad[:20]:
                print(f"  {k}: stored={va} recomputed={vb}")
            return 1
        print(f"CHECK OK: {len(b)} numeric leaves match stored JSON at 1e-6")
        return 0
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(result, indent=2) + "\n")
    print(f"wrote {OUT_PATH}")
    print(f"pooled MASE  model={result['mase_model']}  "
          f"seasonal-naive={result['mase_seasonal_naive']}  "
          f"reconciled={result['pooled']['mase_model_reconciled']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
